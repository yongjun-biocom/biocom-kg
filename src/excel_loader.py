"""
excel_loader.py — 온톨로지_지식그래프.xlsx → nodes.json + edges.json
"""

import json
import os
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "온톨로지_지식그래프.xlsx")
DATA_DIR   = os.path.join(os.path.dirname(__file__), "..", "data")


def sheet_to_dicts(ws) -> list[dict]:
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        rows.append({h: v for h, v in zip(headers, row) if h is not None})
    return rows


def solution_type_to_label(solution_type: str) -> str:
    mapping = {
        "건기식_자사":      "Supplement",
        "건기식_다빈치랩":  "Supplement",
        "식단라인":         "DietLine",
    }
    return mapping.get(solution_type or "", "Supplement")


def build_nodes(wb) -> list[dict]:
    nodes = []

    # N01_검사
    for r in sheet_to_dicts(wb["N01_검사"]):
        nodes.append({**r, "type": "검사", "label": "Exam"})

    # N02_유기산마커
    for r in sheet_to_dicts(wb["N02_유기산마커"]):
        node = {"id": r["id"], "name": r["name_ko"], "type": "유기산마커", "label": "OrganicAcidMarker"}
        for k in ("name_en", "character", "bidirectional", "normal_range", "unit",
                  "high_interpretation", "mechanism", "discordance_notes"):
            if r.get(k):
                node[k] = r[k]
        nodes.append(node)

    # N11_영양소
    for r in sheet_to_dicts(wb["N11_영양소"]):
        node = {**r, "type": "영양소", "label": "Nutrient"}
        nodes.append(node)

    # N12_효소
    for r in sheet_to_dicts(wb["N12_효소"]):
        nodes.append({**r, "type": "효소", "label": "Enzyme"})

    # N13_대사경로
    for r in sheet_to_dicts(wb["N13_대사경로"]):
        nodes.append({**r, "type": "대사경로", "label": "Pathway"})

    # N14_관심사
    for r in sheet_to_dicts(wb["N14_관심사"]):
        nodes.append({**r, "type": "관심사", "label": "Concern"})

    # N15_분류
    for r in sheet_to_dicts(wb["N15_분류"]):
        nodes.append({**r, "type": "분류", "label": "Classification"})

    # N16_솔루션 (건기식 + 식단라인 통합)
    for r in sheet_to_dicts(wb["N16_솔루션"]):
        label = solution_type_to_label(r.get("solution_type"))
        node = {**r, "type": r.get("solution_type", "건기식_자사"), "label": label}
        nodes.append(node)

    # N17_유형레이블
    for r in sheet_to_dicts(wb["N17_유형레이블"]):
        nodes.append({**r, "type": "유형레이블", "label": "TypeLabel"})

    return nodes


def build_edges(wb) -> list[dict]:
    edges = []

    def add_edges(sheet_name, edge_type, src_col="from_id", tgt_col="to_id", extra_cols=None):
        ws = wb[sheet_name]
        for r in sheet_to_dicts(ws):
            src = r.get(src_col)
            tgt = r.get(tgt_col)
            if not src or not tgt:
                continue
            edge = {"type": edge_type, "source": src, "target": tgt}
            if extra_cols:
                for col in extra_cols:
                    if r.get(col):
                        edge[col] = r[col]
            edges.append(edge)

    add_edges("E01_측정",          "E01_측정")
    add_edges("E08_관련효소",      "E08_관련효소")
    add_edges("E10_조효소",        "E10_조효소")
    add_edges("E11_소속",          "E11_소속",   extra_cols=["context"])
    add_edges("E09_필요영양소",    "E09_필요영양소", extra_cols=["basis", "evidence_level"])
    add_edges("E12_관련관심사",    "E12_관련관심사", extra_cols=["context", "evidence_level"])
    add_edges("E13_포함",          "E13_포함",   extra_cols=["amount", "unit"])
    add_edges("E16_추천",          "E16_추천",   extra_cols=["condition"])
    add_edges("E17_억제수단",      "E17_억제수단",
              src_col="marker_id", tgt_col="product_id",
              extra_cols=["pathogen_ref", "marker_name", "product_name"])
    add_edges("E18_제품제약",      "E18_제품제약",
              extra_cols=["constraint_type", "nutrient", "total_amount", "description"])
    add_edges("E19_영양소상호작용","E19_영양소상호작용",
              extra_cols=["interaction_type", "description"])
    add_edges("E02-06_마커간",     "E02_마커간",
              extra_cols=["relation_type", "note"])

    return edges


def run(xlsx_path=None, data_dir=None):
    if xlsx_path is None:
        xlsx_path = XLSX_PATH
    if data_dir is None:
        data_dir = DATA_DIR

    xlsx_path = os.path.abspath(xlsx_path)
    data_dir  = os.path.abspath(data_dir)
    os.makedirs(data_dir, exist_ok=True)

    print(f"읽는 중: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)

    nodes = build_nodes(wb)
    edges = build_edges(wb)

    nodes_path = os.path.join(data_dir, "nodes.json")
    edges_path = os.path.join(data_dir, "edges.json")

    with open(nodes_path, "w", encoding="utf-8") as f:
        json.dump(nodes, f, ensure_ascii=False, indent=2, default=str)
    with open(edges_path, "w", encoding="utf-8") as f:
        json.dump(edges, f, ensure_ascii=False, indent=2, default=str)

    print(f"노드 {len(nodes)}개 → {nodes_path}")
    print(f"엣지 {len(edges)}개 → {edges_path}")

    # 타입별 통계
    from collections import Counter
    nt = Counter(n["type"] for n in nodes)
    et = Counter(e["type"] for e in edges)
    print("노드 타입:", dict(nt))
    print("엣지 타입:", dict(et))

    return nodes, edges


if __name__ == "__main__":
    run()
